import csv
import io
from typing import Optional
import requests
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from database import db
from models import ProductIn
from security import new_id, now_iso, require_business, require_roles

router = APIRouter(tags=["products"])
PRODUCT_FIELDS = {"_id": 0}
MANAGER = Depends(require_roles("propietario", "administrador"))

def _csv_response(rows, headers, filename):
    buf=io.StringIO();buf.write("\ufeff");writer=csv.writer(buf);writer.writerow(headers);writer.writerows(rows);buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]),media_type="text/csv",headers={"Content-Disposition":f"attachment; filename={filename}.csv"})

def _xlsx_response(rows, headers, filename, instructions=False):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    wb=Workbook();ws=wb.active;ws.title="Productos"
    ws.append(headers)
    for row in rows: ws.append(row)
    thin=Side(style="thin",color="D9DEE7")
    for c in ws[1]:
        c.fill=PatternFill("solid",fgColor="1F2937");c.font=Font(color="FFFFFF",bold=True);c.alignment=Alignment(horizontal="center",vertical="center");c.border=Border(bottom=thin)
    for row in ws.iter_rows(min_row=2):
        for c in row: c.border=Border(left=thin,right=thin,top=thin,bottom=thin);c.alignment=Alignment(vertical="top",wrap_text=True)
    ws.freeze_panes="A2";ws.auto_filter.ref=f"A1:{get_column_letter(len(headers))}{max(1,len(rows)+1)}"
    if rows:
        table=Table(displayName="ProductosReporte",ref=f"A1:{get_column_letter(len(headers))}{len(rows)+1}");table.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True);ws.add_table(table)
    for i,h in enumerate(headers,1): ws.column_dimensions[get_column_letter(i)].width=max(12,min(34,max([len(str(h))]+[len(str(r[i-1])) for r in rows[:100]])+3))
    ws.row_dimensions[1].height=28
    if instructions:
        ins=wb.create_sheet("Instrucciones");ins.column_dimensions["A"].width=28;ins.column_dimensions["B"].width=72
        ins.append(["PLANTILLA CUADRAAPP","Carga de productos"]);ins.append(["Importación","Completa la hoja Productos y conserva los encabezados."]);ins.append(["Obligatorio","Nombre del producto."]);ins.append(["Números","Usa números sin símbolos de moneda."]);ins.append(["Códigos","SKU y código de barras se manejan como texto."]);ins["A1"].font=Font(bold=True,size=14);ins["B1"].font=Font(bold=True,size=14)
    buf=io.BytesIO();wb.save(buf);buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]),media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":f'attachment; filename="{filename}.xlsx"'})

def _product_rows(products):
    return [[p.get("name",""),p.get("sku",""),p.get("barcode") or "",p.get("category",""),p.get("brand") or "",p.get("supplier") or "",p.get("purchase_price",0),p.get("sale_price",0),p.get("stock",0),p.get("min_stock",0),p.get("max_stock") if p.get("max_stock") is not None else "",p.get("unit","unidad"),p.get("image_url") or ""] for p in products]

@router.get("/products")
async def list_products(search:Optional[str]=None,category:Optional[str]=None,user:dict=Depends(require_business)):
    query={"business_id":user["business_id"]}
    if category:query["category"]=category
    if search:query["$or"]=[{"name":{"$regex":search,"$options":"i"}},{"sku":{"$regex":search,"$options":"i"}},{"barcode":{"$regex":search,"$options":"i"}}]
    return {"products":await db.products.find(query,PRODUCT_FIELDS).sort("name",1).to_list(5000)}

@router.get("/products/image-search")
async def search_product_images(q:str,user:dict=Depends(require_business)):
    q=q.strip()
    if len(q)<2:raise HTTPException(400,"Indica un producto para buscar")
    try:
        r=requests.get("https://api.openverse.org/v1/images/",params={"q":q,"page_size":12,"mature":"false"},timeout=8);r.raise_for_status();data=r.json();results=[]
        for x in data.get("results",[]):
            url=x.get("thumbnail") or x.get("url")
            if url:results.append({"url":url,"thumbnail":x.get("thumbnail") or url,"title":x.get("title") or q,"source":x.get("foreign_landing_url") or "","creator":x.get("creator") or ""})
        return {"query":q,"images":results}
    except Exception:return {"query":q,"images":[]}

@router.post("/products")
async def create_product(data:ProductIn,user:dict=MANAGER):
    bid=user["business_id"];count=await db.products.count_documents({"business_id":bid});product={"id":new_id(),"business_id":bid,**data.model_dump(),"sku":data.sku or f"P-{count+1:04d}","category":data.category or "General","status":"activo","created_at":now_iso(),"updated_at":now_iso()}
    await db.products.insert_one(product)
    if product["stock"]>0:await db.inventory_movements.insert_one({"id":new_id(),"business_id":bid,"product_id":product["id"],"product_name":product["name"],"type":"entrada","reason":"carga_inicial","quantity":product["stock"],"stock_after":product["stock"],"user_email":user["email"],"notes":None,"created_at":now_iso()})
    product.pop("_id",None);return {"product":product}

@router.put("/products/{product_id}")
async def update_product(product_id:str,data:ProductIn,user:dict=MANAGER):
    bid=user["business_id"];product=await db.products.find_one({"id":product_id,"business_id":bid})
    if not product:raise HTTPException(404,"Producto no encontrado")
    updates=data.model_dump();updates.pop("stock",None);updates["updated_at"]=now_iso();await db.products.update_one({"id":product_id},{"$set":updates})
    return {"product":await db.products.find_one({"id":product_id},PRODUCT_FIELDS)}

@router.delete("/products/{product_id}")
async def delete_product(product_id:str,user:dict=MANAGER):
    result=await db.products.delete_one({"id":product_id,"business_id":user["business_id"]})
    if result.deleted_count==0:raise HTTPException(404,"Producto no encontrado")
    return {"ok":True}

@router.get("/products/export/csv")
async def export_products(user:dict=MANAGER):
    products=await db.products.find({"business_id":user["business_id"]},PRODUCT_FIELDS).sort("name",1).to_list(10000)
    return _csv_response(_product_rows(products),["nombre","sku","codigo_barras","categoria","marca","proveedor","precio_compra","precio_venta","stock","stock_minimo","stock_maximo","unidad","imagen_url"],"productos")

@router.get("/products/export/xlsx")
async def export_products_xlsx(user:dict=MANAGER):
    products=await db.products.find({"business_id":user["business_id"]},PRODUCT_FIELDS).sort("name",1).to_list(10000)
    headers=["Producto","SKU","Código de barras","Categoría","Marca","Proveedor","Costo","Precio","Stock","Stock mínimo","Stock máximo","Unidad","Imagen"]
    return _xlsx_response(_product_rows(products),headers,"productos")

@router.get("/products/template/xlsx")
async def product_template_xlsx(user:dict=MANAGER):
    headers=["nombre","sku","codigo_barras","categoria","marca","proveedor","precio_compra","precio_venta","stock","stock_minimo","stock_maximo","unidad","imagen_url"]
    example=[["Harina P.A.N. blanca 1kg","P-0001","7591234567890","Alimentos","P.A.N.","Proveedor Demo",1.25,1.80,20,5,50,"unidad",""]]
    return _xlsx_response(example,headers,"plantilla_productos",True)

@router.post("/products/import")
async def import_products(file:UploadFile=File(...),user:dict=MANAGER):
    content=(await file.read()).decode("utf-8-sig");reader=csv.DictReader(io.StringIO(content));bid=user["business_id"];count=await db.products.count_documents({"business_id":bid});created=0
    def num(v,default=0.0):
        try:return float(str(v).replace(",",".").strip())
        except (TypeError,ValueError):return default
    for row in reader:
        name=(row.get("nombre") or row.get("name") or "").strip()
        if not name:continue
        count+=1;created+=1;product={"id":new_id(),"business_id":bid,"name":name,"sku":(row.get("sku") or "").strip() or f"P-{count:04d}","barcode":(row.get("codigo_barras") or row.get("barcode") or "").strip() or None,"category":(row.get("categoria") or row.get("category") or "").strip() or "General","brand":(row.get("marca") or "").strip() or None,"supplier":(row.get("proveedor") or "").strip() or None,"purchase_price":num(row.get("precio_compra")),"sale_price":num(row.get("precio_venta")),"stock":num(row.get("stock")),"min_stock":num(row.get("stock_minimo"),5),"max_stock":num(row.get("stock_maximo")) or None,"unit":(row.get("unidad") or "unidad").strip(),"image_url":(row.get("imagen_url") or row.get("image_url") or "").strip() or None,"status":"activo","created_at":now_iso(),"updated_at":now_iso()}
        await db.products.insert_one(product)
        if product["stock"]>0:await db.inventory_movements.insert_one({"id":new_id(),"business_id":bid,"product_id":product["id"],"product_name":product["name"],"type":"entrada","reason":"carga_inicial","quantity":product["stock"],"stock_after":product["stock"],"user_email":user["email"],"notes":"Importado desde CSV","created_at":now_iso()})
    return {"created":created}

@router.post("/products/import/xlsx")
async def import_products_xlsx(file:UploadFile=File(...),user:dict=MANAGER):
    if not (file.filename or "").lower().endswith((".xlsx",".xlsm")):
        raise HTTPException(400,"El archivo debe ser Excel (.xlsx o .xlsm)")
    try:
        from openpyxl import load_workbook
        raw=await file.read();wb=load_workbook(io.BytesIO(raw),read_only=True,data_only=True)
        ws=wb["Productos"] if "Productos" in wb.sheetnames else wb.active
        rows=list(ws.iter_rows(values_only=True))
    except Exception as exc:
        raise HTTPException(400,f"No se pudo leer el Excel: {exc}")
    if not rows: raise HTTPException(400,"El Excel está vacío")
    headers=[str(x).strip().lower() if x is not None else "" for x in rows[0]]
    required="nombre"
    if required not in headers: raise HTTPException(400,"Falta la columna obligatoria 'nombre'")
    idx={h:i for i,h in enumerate(headers) if h}
    bid=user["business_id"]
    existing=await db.products.find({"business_id":bid}, {"_id":0,"sku":1,"barcode":1}).to_list(10000)
    existing_skus={str(x.get("sku")).strip().lower() for x in existing if x.get("sku")}
    existing_barcodes={str(x.get("barcode")).strip() for x in existing if x.get("barcode")}
    created=0;issues=[];seen_skus=set();seen_barcodes=set();count=await db.products.count_documents({"business_id":bid})
    def val(row,key,default=""):
        i=idx.get(key);return row[i] if i is not None and i<len(row) else default
    def text(row,key):
        v=val(row,key,"");return "" if v is None else str(v).strip()
    def num(row,key,default=0.0):
        v=val(row,key,default)
        try:return float(str(v).replace(",",".").strip())
        except (TypeError,ValueError):raise ValueError(f"{key} debe ser numérico")
    for line,row in enumerate(rows[1:],2):
        if not any(x not in (None,"") for x in row): continue
        name=text(row,"nombre")
        if not name: issues.append({"fila":line,"error":"Falta nombre"});continue
        sku=text(row,"sku").lower() or f"p-{count+created+1:04d}"
        barcode=text(row,"codigo_barras")
        if sku in existing_skus or sku in seen_skus: issues.append({"fila":line,"error":f"SKU duplicado: {sku}"});continue
        if barcode and (barcode in existing_barcodes or barcode in seen_barcodes): issues.append({"fila":line,"error":f"Código de barras duplicado: {barcode}"});continue
        try:
            purchase=num(row,"precio_compra");sale=num(row,"precio_venta");stock=num(row,"stock");minimum=num(row,"stock_minimo",5);maximum=num(row,"stock_maximo",0)
        except ValueError as exc: issues.append({"fila":line,"error":str(exc)});continue
        product={"id":new_id(),"business_id":bid,"name":name,"sku":sku,"barcode":barcode or None,"category":text(row,"categoria") or "General","brand":text(row,"marca") or None,"supplier":text(row,"proveedor") or None,"purchase_price":purchase,"sale_price":sale,"stock":stock,"min_stock":minimum,"max_stock":maximum or None,"unit":text(row,"unidad") or "unidad","image_url":text(row,"imagen_url") or None,"status":"activo","created_at":now_iso(),"updated_at":now_iso()}
        await db.products.insert_one(product);created+=1;seen_skus.add(sku);existing_skus.add(sku)
        if barcode: seen_barcodes.add(barcode);existing_barcodes.add(barcode)
        if stock>0: await db.inventory_movements.insert_one({"id":new_id(),"business_id":bid,"product_id":product["id"],"product_name":name,"type":"entrada","reason":"carga_inicial","quantity":stock,"stock_after":stock,"user_email":user["email"],"notes":"Importado desde Excel","created_at":now_iso()})
    return {"created":created,"issues":issues,"errors":len(issues)}
