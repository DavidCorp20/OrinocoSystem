import io
from datetime import datetime, timezone, timedelta
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from database import db
from security import require_roles, now_iso
router = APIRouter(tags=["finance-export"])
MANAGER = Depends(require_roles("propietario", "administrador"))
def _response(wb):
    buf=io.BytesIO();wb.save(buf);buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]),media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":'attachment; filename="finanzas_cuadraapp.xlsx"'})
def _sheet(wb,title,headers,rows):
    ws=wb.create_sheet(title[:31]);ws.append(headers)
    for row in rows:ws.append(row)
    thin=Side(style="thin",color="D9DEE7")
    for c in ws[1]:c.fill=PatternFill("solid",fgColor="1F2937");c.font=Font(color="FFFFFF",bold=True);c.alignment=Alignment(horizontal="center",vertical="center");c.border=Border(bottom=thin)
    for row in ws.iter_rows(min_row=2):
        for c in row:c.border=Border(left=thin,right=thin,top=thin,bottom=thin);c.alignment=Alignment(vertical="top",wrap_text=True)
    ws.freeze_panes="A2"
    if rows:
        ref=f"A1:{get_column_letter(len(headers))}{len(rows)+1}";table=Table(displayName=(title.replace(" ","")+"Reporte")[:25],ref=ref);table.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True);ws.add_table(table)
    for i,h in enumerate(headers,1):
        vals=[len(str(h))]+[len(str(r[i-1])) for r in rows[:100] if i-1<len(r)];ws.column_dimensions[get_column_letter(i)].width=max(12,min(36,max(vals)+3))
    ws.row_dimensions[1].height=28;return ws
@router.get("/finances/export/xlsx")
async def export_finances_xlsx(from_date:str|None=None,to_date:str|None=None,user:dict=MANAGER):
    now=datetime.now(timezone.utc)
    if not from_date:from_date=(now-timedelta(days=30)).date().isoformat()
    if not to_date:to_date=now.date().isoformat()
    q={"business_id":user["business_id"],"created_at":{"$gte":from_date,"$lte":to_date+"T23:59:59"}}
    sales=await db.sales.find(q,{"_id":0}).sort("created_at",-1).to_list(50000);expenses=await db.expenses.find(q,{"_id":0}).sort("created_at",-1).to_list(50000);purchases=await db.purchases.find(q,{"_id":0}).sort("created_at",-1).to_list(50000)
    revenue=round(sum(float(s.get("total",0) or 0) for s in sales),2);gross=round(sum(float(s.get("profit",0) or 0) for s in sales),2);opex=round(sum(float(e.get("amount",0) or 0) for e in expenses),2);purchases_total=round(sum(float(p.get("total",0) or 0) for p in purchases),2);net=round(gross-opex,2);margin=round(gross/revenue*100,1) if revenue else 0
    by_cat={}
    for e in expenses:by_cat[e.get("category","otros")]=round(by_cat.get(e.get("category","otros"),0)+float(e.get("amount",0) or 0),2)
    wb=Workbook();wb.remove(wb.active)
    summary=_sheet(wb,"Resumen financiero",["Indicador","Valor"],[["Período",f"{from_date} a {to_date}"],["Ventas",revenue],["Utilidad bruta",gross],["Gastos operativos",opex],["Compras de mercancía",purchases_total],["Resultado neto estimado",net],["Margen bruto",margin/100],["Cantidad de ventas",len(sales)],["Cantidad de gastos",len(expenses)],["Cantidad de compras",len(purchases)],["Generado",now_iso()]])
    _sheet(wb,"Gastos",["Fecha","Categoría","Descripción","Monto","Método de pago","Usuario"],[[e.get("created_at","")[:10],e.get("category",""),e.get("description",""),e.get("amount",0),e.get("payment_method",""),e.get("user_email","")] for e in expenses])
    _sheet(wb,"Ventas",["Fecha","Factura","Cliente","Total","Costo","Ganancia","Método de pago","Usuario"],[[s.get("created_at","")[:10],s.get("invoice_number",""),s.get("customer_name",""),s.get("total",0),s.get("cost_total",0),s.get("profit",0),s.get("payment_method",""),s.get("user_email","")] for s in sales])
    _sheet(wb,"Compras",["Fecha","Factura proveedor","Proveedor","Total","Base","IVA","Estado","Usuario"],[[p.get("created_at","")[:10],p.get("supplier_invoice_number","") ,p.get("supplier",""),p.get("total",0),p.get("base",0),p.get("iva_amount",0),p.get("status",""),p.get("user_email","")] for p in purchases])
    _sheet(wb,"Gastos por categoría",["Categoría","Monto","% de gastos"],[[cat,amount,amount/opex if opex else 0] for cat,amount in sorted(by_cat.items(),key=lambda x:-x[1])])
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value,(int,float)) and not isinstance(c.value,bool):c.number_format='#,##0.00'
    summary["B8"].number_format="0.0%"
    ws=wb["Gastos por categoría"]
    for c in ws["C"][1:]:c.number_format="0.0%"
    return _response(wb)
