import csv
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from database import db
from security import require_roles, new_id, now_iso

router = APIRouter(tags=["import-export"])
MANAGER = Depends(require_roles("propietario", "administrador"))

PRODUCT_COLUMNS = [
    ("nombre", "Nombre del producto", "Texto", "Harina P.A.N. blanca 1kg"),
    ("sku", "SKU / Código interno", "Texto", "P-0001"),
    ("codigo_barras", "Código de barras", "Texto", "7591234567890"),
    ("categoria", "Categoría", "Texto", "Alimentos"),
    ("marca", "Marca", "Texto", "P.A.N."),
    ("proveedor", "Proveedor", "Texto", "Proveedor Demo"),
    ("precio_compra", "Precio de compra", "Número", "1.25"),
    ("precio_venta", "Precio de venta", "Número", "1.80"),
    ("stock", "Stock inicial", "Número", "20"),
    ("stock_minimo", "Stock mínimo", "Número", "5"),
    ("stock_maximo", "Stock máximo", "Número", "50"),
    ("unidad", "Unidad", "Texto", "unidad"),
    ("imagen_url", "URL de imagen", "Texto", "https://..."),
]


def _csv_response(rows, headers, filename):
    buf = io.StringIO(); buf.write("\ufeff")
    writer = csv.writer(buf); writer.writerow(headers); writer.writerows(rows)
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv", headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'})


def _xlsx_response(wb, filename):
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'})


def _format_sheet(ws, headers, rows, table_name):
    thin = Side(style="thin", color="D9DEE7")
    for col, label in enumerate(headers, 1):
        cell = ws.cell(1, col, label); cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.font = Font(color="FFFFFF", bold=True); cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)
    for r, row in enumerate(rows, 2):
        for c, value in enumerate(row, 1):
            cell = ws.cell(r, c, value); cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows)+1)}"
    if rows:
        ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
        table = Table(displayName=table_name[:25], ref=ref); table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True); ws.add_table(table)
    for i, h in enumerate(headers, 1):
        max_len = max([len(str(h))] + [len(str(r[i-1])) for r in rows[:100] if i-1 < len(r)])
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(34, max_len + 3))
    ws.row_dimensions[1].height = 28


def _add_instructions(wb, columns, title):
    ws = wb.create_sheet("Instrucciones"); ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 70
    ws.append(["PLANTILLA CUADRAAPP", title])
    ws.append(["Uso", "Completa la hoja Productos y conserva exactamente los encabezados."])
    ws.append(["Importación", "CuadraApp valida nombre, SKU y código de barras antes de registrar los productos."])
    ws.append(["Números", "Usa números sin símbolos de moneda. Ejemplo: 125.50"])
    ws.append(["Códigos", "SKU y código de barras se tratan como texto para conservar ceros iniciales."])
    ws.append([]); ws.append(["Columna", "Descripción"])
    for _, label, kind, example in columns: ws.append([label, f"Tipo: {kind}. Ejemplo: {example}"])
    for row in ws.iter_rows():
        for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True); cell.border = Border(bottom=Side(style="thin", color="E5E7EB"))
    ws["A1"].font = Font(bold=True, size=14); ws["B1"].font = Font(bold=True, size=14)


def _num(value, default=0.0):
    try: return float(str(value or "").replace(",", ".").strip())
    except (TypeError, ValueError): return default


@router.get("/templates/products")
async def product_template(user: dict = MANAGER):
    wb = Workbook(); ws = wb.active; ws.title = "Productos"
    headers = [c[1] for c in PRODUCT_COLUMNS]; ws.append(headers); ws.append([c[3] for c in PRODUCT_COLUMNS])
    for c in ws[1]: c.fill = PatternFill("solid", fgColor="1F2937"); c.font = Font(color="FFFFFF", bold=True); c.alignment = Alignment(horizontal="center")
    for c in ws[2]: c.fill = PatternFill("solid", fgColor="EFF6FF")
    ws.freeze_panes = "A2"
    for i, (_, label, *_rest) in enumerate(PRODUCT_COLUMNS, 1): ws.column_dimensions[get_column_letter(i)].width = max(14, min(34, len(label)+4))
    table = Table(displayName="ProductosPlantilla", ref=f"A1:{get_column_letter(len(headers))}2"); table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True); ws.add_table(table)
    _add_instructions(wb, PRODUCT_COLUMNS, "Importación de productos")
    return _xlsx_response(wb, "plantilla_productos")


@router.get("/templates/products/csv")
async def product_template_csv(user: dict = MANAGER): return _csv_response([[c[3] for c in PRODUCT_COLUMNS]], [c[0] for c in PRODUCT_COLUMNS], "plantilla_productos")


@router.post("/products/import/xlsx")
async def import_products_xlsx(file: UploadFile = File(...), user: dict = MANAGER):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")): raise HTTPException(400, "Sube la plantilla Excel .xlsx")
    try:
        wb = load_workbook(io.BytesIO(await file.read()), read_only=True, data_only=True); ws = wb["Productos"] if "Productos" in wb.sheetnames else wb.active; rows = list(ws.iter_rows(values_only=True))
    except Exception as exc: raise HTTPException(400, f"No pudimos leer el Excel: {exc}")
    if not rows: raise HTTPException(400, "El archivo está vacío")
    headers = [str(x or "").strip().lower() for x in rows[0]]
    if "nombre" not in headers: raise HTTPException(400, "Falta la columna obligatoria: nombre")
    bid = user["business_id"]; existing = await db.products.find({"business_id": bid}, {"_id": 0, "sku": 1, "barcode": 1}).to_list(20000)
    skus = {str(p.get("sku")) for p in existing if p.get("sku")}; barcodes = {str(p.get("barcode")) for p in existing if p.get("barcode")}; created = errors = 0; issues = []; base_count = await db.products.count_documents({"business_id": bid})
    for line_no, values in enumerate(rows[1:], 2):
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}; name = str(row.get("nombre") or "").strip()
        if not name or name.startswith("Ej."): continue
        sku = str(row.get("sku") or "").strip() or f"P-{base_count+created+1:04d}"; barcode = str(row.get("codigo_barras") or "").strip() or None
        if sku in skus: errors += 1; issues.append({"row": line_no, "field": "sku", "message": "SKU ya existe"}); continue
        if barcode and barcode in barcodes: errors += 1; issues.append({"row": line_no, "field": "codigo_barras", "message": "Código de barras ya existe"}); continue
        product = {"id":new_id(),"business_id":bid,"name":name,"sku":sku,"barcode":barcode,"category":str(row.get("categoria") or "General").strip() or "General","brand":str(row.get("marca") or "").strip() or None,"supplier":str(row.get("proveedor") or "").strip() or None,"purchase_price":_num(row.get("precio_compra")),"sale_price":_num(row.get("precio_venta")),"stock":_num(row.get("stock")),"min_stock":_num(row.get("stock_minimo"),5),"max_stock":_num(row.get("stock_maximo")) or None,"unit":str(row.get("unidad") or "unidad").strip(),"base_unit":str(row.get("unidad") or "unidad").strip(),"image_url":str(row.get("imagen_url") or "").strip() or None,"status":"activo","created_at":now_iso(),"updated_at":now_iso()}
        await db.products.insert_one(product); skus.add(sku); created += 1
        if barcode: barcodes.add(barcode)
        if product["stock"] > 0: await db.inventory_movements.insert_one({"id":new_id(),"business_id":bid,"product_id":product["id"],"product_name":name,"type":"entrada","reason":"carga_inicial","quantity":product["stock"],"stock_after":product["stock"],"user_email":user["email"],"notes":"Importado desde plantilla Excel","created_at":now_iso()})
    return {"created":created,"errors":errors,"issues":issues,"processed":created+errors}


async def _dataset(resource, bid, from_date=None, to_date=None):
    q = {"business_id": bid}
    if resource in {"ventas","compras","gastos"} and from_date: q["created_at"] = {"$gte": from_date}
    if resource in {"ventas","compras","gastos"} and to_date: q.setdefault("created_at", {})["$lte"] = to_date + "T23:59:59"
    if resource == "productos":
        docs = await db.products.find(q,{"_id":0}).sort("name",1).to_list(50000); return [[p.get("name",""),p.get("sku",""),p.get("barcode",""),p.get("category",""),p.get("brand",""),p.get("supplier",""),p.get("purchase_price",0),p.get("sale_price",0),p.get("stock",0),p.get("min_stock",0),p.get("max_stock", ""),p.get("unit","unidad"),p.get("status","")] for p in docs],["Producto","SKU","Código de barras","Categoría","Marca","Proveedor","Costo","Precio","Stock","Stock mínimo","Stock máximo","Unidad","Estado"]
    if resource == "ventas":
        docs = await db.sales.find(q,{"_id":0}).sort("created_at",-1).to_list(50000); return [[s.get("created_at","")[:10],s.get("invoice_number",""),"; ".join(f"{i.get('name','')} x{i.get('quantity',0):g}" for i in s.get("items",[])),s.get("payment_method",""),s.get("customer_name",""),s.get("total",0),s.get("cost_total",0),s.get("profit",0),s.get("user_email","")] for s in docs],["Fecha","Factura","Productos","Método de pago","Cliente","Total","Costo","Ganancia","Usuario"]
    if resource == "compras":
        docs = await db.purchases.find(q,{"_id":0}).sort("created_at",-1).to_list(50000); return [[p.get("created_at","")[:10],p.get("invoice_number",""),p.get("supplier",""),"; ".join(f"{i.get('name','')} x{i.get('quantity',0):g}" for i in p.get("items",[])),p.get("status",""),p.get("total",0),p.get("base",0),p.get("iva_amount",0),p.get("user_email","")] for p in docs],["Fecha","Comprobante","Proveedor","Productos","Estado","Total","Base","IVA","Usuario"]
    if resource == "movimientos":
        docs = await db.inventory_movements.find(q,{"_id":0}).sort("created_at",-1).to_list(50000); return [[m.get("created_at","")[:10],m.get("product_name",""),m.get("type",""),m.get("reason",""),m.get("quantity",0),m.get("stock_after",""),m.get("user_email",""),m.get("notes","")] for m in docs],["Fecha","Producto","Tipo","Motivo","Cantidad","Stock resultante","Usuario","Notas"]
    if resource == "gastos":
        docs = await db.expenses.find(q,{"_id":0}).sort("created_at",-1).to_list(50000); return [[e.get("created_at","")[:10],e.get("category",""),e.get("description",""),e.get("amount",0),e.get("payment_method",""),e.get("user_email","")] for e in docs],["Fecha","Categoría","Descripción","Monto","Método de pago","Usuario"]
    raise HTTPException(404,"Reporte no disponible")


@router.get("/exports/{resource}/xlsx")
async def export_resource_xlsx(resource: str, from_date: str | None = None, to_date: str | None = None, user: dict = MANAGER):
    if resource not in {"productos","ventas","compras","movimientos","gastos"}: raise HTTPException(404,"Reporte no disponible")
    rows, headers = await _dataset(resource,user["business_id"],from_date,to_date)
    wb = Workbook(); ws = wb.active; ws.title = resource.title()[:31]; _format_sheet(ws,headers,rows,resource.title())
    summary = wb.create_sheet("Resumen"); summary.append(["Reporte CuadraApp",resource.title()]); summary.append(["Registros",len(rows)]); summary.append(["Desde",from_date or "Todos"]); summary.append(["Hasta",to_date or "Todos"]); summary.append(["Generado",now_iso()]); summary.column_dimensions["A"].width=24; summary.column_dimensions["B"].width=36
    for c in summary[1]: c.font=Font(bold=True,size=14)
    return _xlsx_response(wb,f"reporte_{resource}")
